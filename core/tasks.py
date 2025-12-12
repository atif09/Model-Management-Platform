import csv
import json
import logging
from io import StringIO
from django.core.files.storage import default_storage
from celery import shared_task
from core.models import Job, Dataset

logger = logging.getLogger(__name__)


# ============ CSV VALIDATION TASK ============
@shared_task(bind=True, max_retries=3)
def validate_csv(self, dataset_id, job_id):
    """
    Validate CSV file and count rows.
    Updates Job progress from 0-100%.
    
    Retry logic: waits 3s, then 9s, then 27s before giving up
    """
    try:
        # Get the Job and Dataset from database
        job = Job.objects.get(id=job_id)
        dataset = Dataset.objects.get(id=dataset_id)
        
        # Mark job as PROCESSING and store the Celery task ID
        job.status = 'PROCESSING'
        job.task_id = self.request.id  # Store task ID for tracking
        job.save()
        
        # Read the CSV file from storage
        file_path = dataset.file.name
        file_content = default_storage.open(file_path, 'r').read()
        
        # 25% - File loaded
        job.progress = 25
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 25})
        
        # Parse CSV and convert to list of rows
        csv_reader = csv.DictReader(StringIO(file_content))
        rows = list(csv_reader)
        
        # 75% - Parsed successfully
        job.progress = 75
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 75})
        
        # Get column headers from CSV
        headers = csv_reader.fieldnames or []
        
        # Create validation report
        validation_report = {
            'is_valid': len(rows) > 0,
            'row_count': len(rows),
            'headers': headers,
            'header_count': len(headers),
        }
        
        # 100% - COMPLETE
        job.progress = 100
        job.status = 'COMPLETED'
        job.result_data = validation_report
        job.save()
        
        logger.info(f"CSV validation completed for job {job_id}: {len(rows)} rows")
        return validation_report
        
    except Exception as exc:
        # Something went wrong - log it and mark job as FAILED
        logger.error(f"CSV validation failed for job {job_id}: {str(exc)}")
        job = Job.objects.get(id=job_id)
        job.status = 'FAILED'
        job.error_message = str(exc)
        job.save()
        
        # Retry with exponential backoff: 3^1=3s, 3^2=9s, 3^3=27s
        raise self.retry(exc=exc, countdown=3 ** self.request.retries)

  


# ============ IMAGE PROCESSING TASK ============
@shared_task(bind=True, max_retries=3)
def process_image(self, dataset_id, job_id):
    """
    Process image: resize and generate thumbnail.
    Updates Job progress from 0-100%.
    """
    try:
        from PIL import Image
        
        job = Job.objects.get(id=job_id)
        dataset = Dataset.objects.get(id=dataset_id)
        
        job.status = 'PROCESSING'
        job.task_id = self.request.id
        job.save()
        
        # Open original image file
        file_path = dataset.file.name
        image = Image.open(default_storage.open(file_path, 'rb'))
        
        # 30% - Image loaded
        job.progress = 30
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 30})
        
        # Get original dimensions before any changes
        original_size = image.size
        
        # 50% - Creating resized version
        job.progress = 50
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 50})
        
        # Create a resized copy (800x600 max)
        resized = image.copy()
        resized.thumbnail((800, 600), Image.Resampling.LANCZOS)
        resized_path = f"processed/{dataset.id}_resized.jpg"
        default_storage.save(resized_path, resized)
        
        # 75% - Creating thumbnail
        job.progress = 75
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 75})
        
        # Create a thumbnail (200x200 max)
        thumb = image.copy()
        thumb.thumbnail((200, 200), Image.Resampling.LANCZOS)
        thumb_path = f"processed/{dataset.id}_thumb.jpg"
        default_storage.save(thumb_path, thumb)
        
        # 100% - COMPLETE
        job.progress = 100
        job.status = 'COMPLETED'
        job.result_data = {
            'original_size': original_size,
            'resized_path': resized_path,
            'thumbnail_path': thumb_path,
            'format': image.format,
        }
        job.save()
        
        logger.info(f"Image processing completed for job {job_id}")
        return job.result_data
        
    except Exception as exc:
        logger.error(f"Image processing failed for job {job_id}: {str(exc)}")
        job = Job.objects.get(id=job_id)
        job.status = 'FAILED'
        job.error_message = str(exc)
        job.save()
        raise self.retry(exc=exc, countdown=3 ** self.request.retries)

  


# ============ STATISTICS GENERATION TASK ============
@shared_task(bind=True, max_retries=3)
def generate_statistics(self, dataset_id, job_id):
    """
    Generate dataset statistics and quality metrics.
    Updates Job progress from 0-100%.
    """
    try:
        job = Job.objects.get(id=job_id)
        dataset = Dataset.objects.get(id=dataset_id)
        
        job.status = 'PROCESSING'
        job.task_id = self.request.id
        job.save()
        
        # Read CSV file
        file_path = dataset.file.name
        file_content = default_storage.open(file_path, 'r').read()
        
        # 25% - File loaded
        job.progress = 25
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 25})
        
        # Parse CSV into rows
        csv_reader = csv.DictReader(StringIO(file_content))
        rows = list(csv_reader)
        
        # 50% - Parsed
        job.progress = 50
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 50})
        
        # Get column names
        headers = csv_reader.fieldnames or []
        
        # Calculate statistics
        stats = {
            'total_rows': len(rows),
            'total_columns': len(headers),
            'column_names': headers,
            'null_counts': {},
        }
        
        # Count missing/null values per column
        for header in headers:
            null_count = sum(1 for row in rows if not row.get(header))
            stats['null_counts'][header] = null_count
        
        # 75% - Analyzed
        job.progress = 75
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 75})
        
        # 100% - COMPLETE
        job.progress = 100
        job.status = 'COMPLETED'
        job.result_data = stats
        job.save()
        
        logger.info(f"Statistics generated for job {job_id}: {len(rows)} rows analyzed")
        return stats
        
    except Exception as exc:
        logger.error(f"Statistics generation failed for job {job_id}: {str(exc)}")
        job = Job.objects.get(id=job_id)
        job.status = 'FAILED'
        job.error_message = str(exc)
        job.save()
        raise self.retry(exc=exc, countdown=3 ** self.request.retries)



# ============ FILE FORMAT CONVERSION TASK ============
@shared_task(bind=True, max_retries=3)
def convert_file_format(self, dataset_id, job_id, target_format):
    """
    Convert file to target format (json, excel).
    Updates Job progress from 0-100%.
    
    target_format: 'json' or 'excel'
    """
    try:
        job = Job.objects.get(id=job_id)
        dataset = Dataset.objects.get(id=dataset_id)
        
        job.status = 'PROCESSING'
        job.task_id = self.request.id
        job.save()
        
        # Read CSV file
        file_path = dataset.file.name
        file_content = default_storage.open(file_path, 'r').read()
        
        # 25% - File loaded
        job.progress = 25
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 25})
        
        # Parse CSV into rows
        csv_reader = csv.DictReader(StringIO(file_content))
        rows = list(csv_reader)
        headers = csv_reader.fieldnames or []
        
        # 50% - Parsed
        job.progress = 50
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 50})
        
        # Convert based on target format
        if target_format == 'json':
            # Convert to JSON format
            json_content = json.dumps(rows, indent=2)
            output_path = f"converted/{dataset.id}_converted.json"
            default_storage.save(output_path, StringIO(json_content))
        
        elif target_format == 'excel':
            # Convert to Excel format
            try:
                import openpyxl
            except ImportError:
                raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write headers in first row
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
            
            # Write data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row_idx, col_idx, row.get(header))
            
            output_path = f"converted/{dataset.id}_converted.xlsx"
            wb.save(default_storage.open(output_path, 'wb'))
        
        # 75% - Converted
        job.progress = 75
        job.save()
        self.update_state(state='PROGRESS', meta={'progress': 75})
        
        # 100% - COMPLETE
        job.progress = 100
        job.status = 'COMPLETED'
        job.result_data = {
            'original_format': 'csv',
            'target_format': target_format,
            'output_path': output_path,
            'row_count': len(rows),
        }
        job.save()
        
        logger.info(f"File conversion completed for job {job_id}: {target_format}")
        return job.result_data
        
    except Exception as exc:
        logger.error(f"File conversion failed for job {job_id}: {str(exc)}")
        job = Job.objects.get(id=job_id)
        job.status = 'FAILED'
        job.error_message = str(exc)
        job.save()
        raise self.retry(exc=exc, countdown=3 ** self.request.retries)

  